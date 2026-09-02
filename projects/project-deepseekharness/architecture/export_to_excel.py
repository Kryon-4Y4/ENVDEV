#!/usr/bin/env python3
"""将 deepseek-harness-architecture 的 JSON 规范导出为多 Sheet Excel 表格"""

import json
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ── 读取源数据 ──────────────────────────────────────────────
src = Path(__file__).resolve().parent / "dsh-arch-candidate.json"
data = json.loads(src.read_text(encoding="utf-8"))

out = src.parent / "deepseek-harness-architecture.xlsx"

# ── 样式 ────────────────────────────────────────────────────
header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill("solid", fgColor="2F5496")
header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
wrap_align = Alignment(vertical="top", wrap_text=True)

def style_header(ws, ncols):
    for col in range(1, ncols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

def auto_width(ws, min_w=10, max_w=40):
    for col_cells in ws.columns:
        length = min_w
        col_letter = col_cells[0].column_letter
        for cell in col_cells:
            if cell.value:
                length = max(length, min(len(str(cell.value)) + 2, max_w))
        ws.column_dimensions[col_letter].width = length

# ── 创建工作簿 ──────────────────────────────────────────────
wb = Workbook()

# ═══════════════════════════════════════════════════════════
# Sheet 1: 概览 (Meta)
# ═══════════════════════════════════════════════════════════
ws_meta = wb.active
ws_meta.title = "概览"
meta = data["meta"]
ws_meta.append(["属性", "值"])
ws_meta.append(["标题", meta["title"]])
ws_meta.append(["语言", meta["locale"]])
ws_meta.append(["质量级别", meta["quality_profile"]])
ws_meta.append(["画布尺寸", f"{meta['viewBox'][0]} × {meta['viewBox'][1]}"])
ws_meta.append(["组件总数", len(data["components"])])
ws_meta.append(["边界总数", len(data["boundaries"])])
ws_meta.append(["连线总数", len(data["connections"])])
ws_meta.append(["视图总数", len(meta["views"])])
ws_meta.append(["卡片总数", len(data["cards"])])
style_header(ws_meta, 2)
auto_width(ws_meta)

# ═══════════════════════════════════════════════════════════
# Sheet 2: 组件 (Components)
# ═══════════════════════════════════════════════════════════
ws_comp = wb.create_sheet("组件")
ws_comp.append(["ID", "类型", "标签", "副标签", "X", "Y", "宽", "高", "标签"])
for c in data["components"]:
    ws_comp.append([
        c["id"], c["type"], c["label"],
        c.get("sublabel", ""),
        c["pos"][0], c["pos"][1],
        c["size"][0], c["size"][1],
        c.get("tag", ""),
    ])
style_header(ws_comp, 9)
auto_width(ws_comp)

# ═══════════════════════════════════════════════════════════
# Sheet 3: 边界 (Boundaries)
# ═══════════════════════════════════════════════════════════
ws_bnd = wb.create_sheet("边界")
ws_bnd.append(["序号", "类型", "标签", "包含组件"])
for i, b in enumerate(data["boundaries"], 1):
    ws_bnd.append([i, b["kind"], b["label"], ", ".join(b["wraps"])])
style_header(ws_bnd, 4)
auto_width(ws_bnd, max_w=60)

# ═══════════════════════════════════════════════════════════
# Sheet 4: 连线 (Connections)
# ═══════════════════════════════════════════════════════════
ws_conn = wb.create_sheet("连线")
ws_conn.append(["ID", "起始组件", "目标组件", "起始方向", "目标方向", "样式", "路径点"])
for c in data["connections"]:
    via_str = ""
    if "via" in c:
        via_str = " → ".join([f"({p[0]},{p[1]})" for p in c["via"]])
    ws_conn.append([
        c["id"], c["from"], c["to"],
        c.get("fromSide", ""), c.get("toSide", ""),
        c.get("variant", "solid"),
        via_str,
    ])
style_header(ws_conn, 7)
auto_width(ws_conn)

# ═══════════════════════════════════════════════════════════
# Sheet 5: 视图 (Views)
# ═══════════════════════════════════════════════════════════
ws_view = wb.create_sheet("视图")
ws_view.append(["视图ID", "标签", "聚焦组件", "说明"])
for v in meta["views"]:
    ws_view.append([v["id"], v["label"], ", ".join(v["focus"]), v["note"]])
style_header(ws_view, 4)
auto_width(ws_view, max_w=60)

# ═══════════════════════════════════════════════════════════
# Sheet 6: 卡片 (Cards) — 架构说明摘要
# ═══════════════════════════════════════════════════════════
ws_card = wb.create_sheet("架构卡片")
ws_card.append(["颜色", "标题", "说明条目"])
for card in data["cards"]:
    for item in card["items"]:
        ws_card.append([card["dot"], card["title"], item])
style_header(ws_card, 3)
auto_width(ws_card, max_w=80)

# ═══════════════════════════════════════════════════════════
# Sheet 7: 组件-边界映射 (交叉矩阵)
# ═══════════════════════════════════════════════════════════
ws_matrix = wb.create_sheet("组件-边界映射")
boundaries = data["boundaries"]
components = data["components"]
comp_ids = [c["id"] for c in components]
bnd_labels = [b["label"] for b in boundaries]

ws_matrix.append(["组件 \\ 边界"] + bnd_labels)
for c in components:
    row = [c["label"]]
    for b in boundaries:
        row.append("✓" if c["id"] in b["wraps"] else "")
    ws_matrix.append(row)
style_header(ws_matrix, len(bnd_labels) + 1)
auto_width(ws_matrix)

# ── 保存 ────────────────────────────────────────────────────
wb.save(out)
print(f"✅ 已导出: {out}")
print(f"   共 {len(wb.sheetnames)} 个工作表: {', '.join(wb.sheetnames)}")
